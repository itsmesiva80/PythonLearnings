import datetime
from datetime import timedelta
from openpyxl import Workbook, load_workbook
wb = load_workbook('/users/sthammana/Documents/CI_Jenkis_List.xlsx')
ws = wb.active
for i in range(2,86):
	cellVal = str(ws.cell(row=i,column = 3).value)
	cellVal1 = ws.cell(row=i,column = 3).value
	cellVal2 = ws.cell(row=i,column = 4).value
	if cellVal != "None":
		timeDiff = datetime.datetime.strptime(cellVal1, "%Y-%m-%d %H:%M")-datetime.datetime.strptime(cellVal2, "%Y-%m-%d %H:%M")
		minDiff=timeDiff/timedelta(minutes=1)
		minDiff=str(minDiff).replace(".0"," minutes").replace("-","")
		ws.cell(row=i,column = 5).value = minDiff
		print(minDiff)
wb.save('/users/sthammana/Documents/CI_Jenkis_List.xlsx')
