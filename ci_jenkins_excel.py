import datetime
from openpyxl import Workbook, load_workbook
wb = load_workbook('/users/sthammana/Documents/CI_Jenkis_List.xlsx')
#wb = Workbook()
ws = wb.active
for i in range(2,86):
	for j in range(3,5):
		cellVal = str(ws.cell(row=i,column = j).value)
		#val1 = cellVal.replace("PM","")
		#val2 = val1.replace("AM","")
		if cellVal != "None":
			val = cellVal.replace("at ","")
			val1 = val.replace(",","")
			print(val1)
		#val5=parser.parse(val4).isoformat()
		#print(val5)
			do = datetime.datetime.strptime(val1, "%b %d %Y %I:%M %p")
			formatDate = do.strftime("%Y-%m-%d %H:%M")
			ws.cell(row=i,column = j).value = formatDate
wb.save('/users/sthammana/Documents/CI_Jenkis_List.xlsx')
